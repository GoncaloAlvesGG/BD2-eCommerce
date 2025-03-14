from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib.auth import logout
from .models import *
from django.contrib.auth.forms import UserChangeForm
from django.db import IntegrityError, connection
from django.http import Http404, JsonResponse
from django.contrib import messages
from collections import defaultdict
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
import json, random, bcrypt, pandas as pd
from .models import EncomendaView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from pymongo import MongoClient
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings 
import string
import random
from datetime import datetime


# Conectar ao MongoDB
client = MongoClient("mongodb://localhost:27017/")
db = client["Loja_online"]
wishlist_collection = db["wishlist"]

#Remove one unit from Cart
def remove_one_from_cart(request, produto_id):
    """Remove 1 unidade de um produto do carrinho. Se a quantidade for 1, remove o produto completamente."""
    
    # Verifica se o carrinho existe na sessão
    if 'cart' not in request.session or not request.session['cart']:
        return JsonResponse({'error': 'Carrinho vazio'}, status=400)
    
    cart = request.session['cart']
    produto_id = str(produto_id)
    
    # Verifica se o produto está no carrinho
    if produto_id not in cart:
        return JsonResponse({'error': 'Produto não encontrado no carrinho'}, status=404)
    
    # Diminui a quantidade em 1
    if cart[produto_id]['quantity'] > 1:
        cart[produto_id]['quantity'] -= 1
    else:
        del cart[produto_id]  # Remove completamente se a quantidade for 1
    
    request.session.modified = True #Save

    return JsonResponse({"message": "Produto atualizado no carrinho!", "cart": cart}, status=200)


def hash_password(password):
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def check_password(password, hashed_password):
    return bcrypt.checkpw(password.encode("utf-8"), hashed_password.encode("utf-8"))

#WISHLIST
def add_to_wishlist(request, produto_id):
    """ Adiciona um produto à wishlist do usuário baseado no nome armazenado na sessão """
    if request.method == "POST":
        try:
            # Verifica se o nome do usuário está na sessão
            owner = request.session.get("nome")  # Obtém o nome da sessão
            
            if not owner:
                return JsonResponse({"error": "Usuário não autenticado!"}, status=403)

            data = json.loads(request.body)
            nome = data.get("nome")

            # Verifica se o produto já está na wishlist do usuário
            if wishlist_collection.find_one({"produto_id": produto_id, "owner": owner}):
                return JsonResponse({"message": "Produto já está na sua wishlist!"}, status=400)

            # Adiciona ao MongoDB com o owner baseado na sessão
            wishlist_collection.insert_one({
                "produto_id": produto_id,
                "nome": nome,
                "owner": owner
            })

            return JsonResponse({"message": "Produto adicionado à sua wishlist!"}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

def wishlist(request):
    """ Busca apenas os produtos da wishlist do usuário logado """
    owner = request.session.get("nome")  # Obtém o nome da sessão

    if not owner:
        return render(request, 'login.html', {'login_error': 'Por favor, faça login para acessar a wishlist.'})

    # Filtra apenas os produtos do usuário autenticado
    items = list(wishlist_collection.find({"owner": owner}, {"_id": 0}))
    produto_ids = tuple(produto["produto_id"] for produto in items)
    produtos_detalhados = []

    with connection.cursor() as cursor:
        for produto_id in produto_ids:
            cursor.execute("SELECT * FROM sp_Produto_READ(%s)", [produto_id])
            colunas = [col[0] for col in cursor.description]  # Obter nomes das colunas
            for row in cursor.fetchall():
                produtos_detalhados.append(dict(zip(colunas, row)))


    return render(request, "wishlist.html", {"items": produtos_detalhados})


def remove_from_wishlist(request, produto_id):
    wishlist_collection.delete_one({"produto_id": produto_id})
    return JsonResponse({"message": "Produto removido da wishlist!"})



def produto_detalhe(request, produto_id):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM sp_Produto_READ(%s)", [produto_id])
        columns = [col[0] for col in cursor.description]
        result = cursor.fetchone()
        cursor.execute("SELECT * FROM recomendar_produtos(%s)", [produto_id])
        colunas = [col[0] for col in cursor.description] 
        resultados = [dict(zip(colunas, linha)) for linha in cursor.fetchall()] 
    
    # Se o produto não for encontrado, levantar erro 404
    if not result:
        raise Http404("Produto não encontrado")
    
    # Formatar o resultado como dicionário
    produto = dict(zip(columns, result))
    
    # Renderizar o template e passar os dados
    return render(request, 'produto_detalhe.html', {'produto': produto, 'recomendacoes': resultados})

def recomendar_produtos_user(user_id):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM recomendar_produtos_user(%s);", [user_id])
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return results

def encomenda(request, encomenda_id):
    utilizador_id = request.session.get('utilizador_id')
    is_admin = request.session.get('is_admin', False)

    encomenda_list = obter_encomenda(encomenda_id)  # Supondo que retorna uma lista
    encomenda = encomenda_list[0] if encomenda_list else None

    # Se não houver encomenda ou o utilizador não for o dono nem admin, negar acesso
    if not encomenda or (not is_admin and encomenda['utilizador_id'] != utilizador_id):
        return render(request, 'erro.html', {'mensagem': 'Não tem permissão para ver esta encomenda.'})

    return render(request, 'detalhe_encomenda.html', {'encomenda': encomenda})


def admin_dashboard(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT estado, total, ultima_encomenda FROM total_encomendas_por_estado")
        estados_encomenda = cursor.fetchall()

    total_pendentes = 0
    total_enviadas_ultima_semana = 0
    ultima_encomenda = None

    for estado in estados_encomenda:
        if estado[0] == 'pendente':
            total_pendentes = estado[1]
        elif estado[0] == 'enviada':
            total_enviadas_ultima_semana = estado[1]
        ultima_encomenda = estado[2]  # A última encomenda será atribuída para qualquer estado

    return render(request, 'dashboard.html', {
        'total_pendentes': total_pendentes,
        'total_enviadas_ultima_semana': total_enviadas_ultima_semana,
        'ultima_encomenda': ultima_encomenda,
    })

def obter_faturas_fornecedor(request, fornecedor_id):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM obter_faturas_fornecedor(%s)", [fornecedor_id])
        rows = cursor.fetchall()
    
    faturas = []
    for row in rows:
        # Formatar a data no formato "H:i - d/m/Y"
        data_emissao = row[2]
        if data_emissao:
            formatted_date = data_emissao.strftime('%H:%M - %d/%m/%Y')
        else:
            formatted_date = None
        
        faturas.append({
            'id_fatura': row[0],
            'data_emissao': formatted_date,
            'valor_total': row[3]
        })
    
    return JsonResponse({'faturas': faturas})



def registo(request):
    storage = messages.get_messages(request)
    storage.used = True  # Marca todas como usadas
    
    if request.method == "POST":
        nome = request.POST['nome']
        email = request.POST['email']
        password = request.POST['password']
        isAdmin = False  # Supondo que a flag seja False por padrão

        # Gerar o hash da senha com bcrypt
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        try:
            # Chamar a stored procedure
            with connection.cursor() as cursor:
                cursor.callproc('sp_Utilizador_CREATE', [nome, email, password_hash, isAdmin])
            
            messages.success(request, "Utilizador registado com sucesso!")
            return redirect('login')

        except Exception as e:
            # Extrair apenas a mensagem de erro relevante (antes de "CONTEXT")
            error_message = str(e).split('CONTEXT:')[0].strip()
            messages.error(request, error_message)  # Exibir a mensagem de erro no frontend

    return render(request, 'registo.html')

def login(request):
    if request.method == "POST":
        email = request.POST['email'].strip().lower()
        password = request.POST['password']

        with connection.cursor() as cursor:
            cursor.callproc('VerificarLogin', [email])
            result = cursor.fetchone()

        categorias = get_all_categories()

        if result is None:
            return render(request, 'login.html', {'login_error': 'Email ou senha inválidos.'})

        utilizador_id, nome, email, stored_password_hash, is_admin = result

        if check_password(password, stored_password_hash):
            request.session['utilizador_id'] = utilizador_id
            request.session['nome'] = nome
            request.session['email'] = email
            request.session['is_admin'] = is_admin
            request.session['categorias'] = categorias

            # Verifica se este utilizador deve redefinir a senha
            if request.session.get("forcar_redefinir_senha") and request.session.get("utilizador_id_reset") == utilizador_id:
                return redirect('alterar_senha')

            messages.success(request, f"Bem-vindo(a), {nome}!")
            return redirect('index')

        return render(request, 'login.html', {'login_error': 'Email ou senha inválidos.'})

    return render(request, 'login.html')


def logout_view(request):
    # Página inicial com a lista de produtos
    logout(request)
    return render(request, "login.html")

def index(request):

    produtos = produtos_4recentes()
    produtos_stock = produtos_mais_stock()
    categorias = get_all_categories()
    utilizador_id = request.session.get('utilizador_id')
    produtos_user = recomendar_produtos_user(utilizador_id)
    request.session['categorias'] = categorias
    return render(request, 'index.html', {'produtos': produtos, 'produtos_stock': produtos_stock, 'produtos_user': produtos_user})

def procurar_produto(request):
    titulo = 'Resultados da sua pesquisa'
    texto = request.GET['nome']
    produtos = procurar_produto_nome(texto)
    return render(request, 'mostrar_produtos.html', {'produtos': produtos, "titulo": titulo})

def mostrar_todos_produtos(request):
    titulo = 'todos'
    produtos = get_all_produtos()
    return render(request, 'mostrar_produtos.html', {'produtos': produtos, "titulo": titulo})

def produtos_categoria(request, categoria_id):
    titulo = 'Produtos da categoria:'
    produtos = obter_produtos_categoria(categoria_id)
    with connection.cursor() as cursor:
        # Call the stored function sp_Produto_READ with produto_id
        cursor.execute("SELECT * FROM sp_Categoria_READ(%s);", [categoria_id])
        
        # Fetch the result (this function returns a table, so fetchone is appropriate for a single product)
        result = cursor.fetchone()
        
        # If result is None, no product was found
        if result:
            categoria_data = {
                'categoria_id': result[0],
                'nome': result[1],
            }
            titulo = "Produtos da categoria: " + categoria_data['nome']
        else:
            return None
    return render(request, 'mostrar_produtos.html', {'produtos': produtos, "titulo": titulo})

def checkout(request):
    # Página de checkout (vazia por enquanto)
    return render(request, 'checkout.html')

# @login_required
def perfil(request):
        # O utilizador está autenticado
        utilizador_id = request.session.get('utilizador_id')
        encomendas = obter_encomendas_utilizador(utilizador_id)
        form = UserChangeForm()
    # Passar os dados para o template
        return render(request, 'perfil.html', {'form': form, 'encomendas': encomendas})

def dashboard_encomendas(request):
    encomendas = obter_encomendas()
    return render(request, 'dashboard_encomendas.html', {'encomendas': encomendas})

def filtrar_encomendas(request):
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    encomendas = obter_encomendas_por_data(data_inicio, data_fim)
    return render(request, 'dashboard_encomendas.html', {'encomendas': encomendas})

def dashboard_produtos(request):
    produtos = ProdutoView.objects.all()
    categorias = get_categorias() 
    fornecedores = get_fornecedores()

    categoria_dict = {categoria['categoria_id']: categoria['nome'] for categoria in categorias}

    # Add the category name to each produto
    for produto in produtos:
        # Look up the category name based on categoria_id
        categoria_nome = categoria_dict.get(produto.categoria_id, 'Categoria desconhecida')
        # Add the category name to the produto object
        produto.categoria_nome = categoria_nome
    
    return render(request, 'dashboard_produtos.html', {'produtos': produtos, 'categorias': categorias, 'fornecedores': fornecedores, 'categorias_json': json.dumps(categorias)})

def get_product_data(produto_id):
    with connection.cursor() as cursor:
        # Call the stored function sp_Produto_READ with produto_id
        cursor.execute("SELECT * FROM sp_Produto_READ(%s);", [produto_id])
        
        # Fetch the result (this function returns a table, so fetchone is appropriate for a single product)
        result = cursor.fetchone()
        
        # If result is None, no product was found
        if result:
            produto_data = {
                'produto_id': result[0],
                'nome': result[1],
                'descricao': result[2],
                'preco': result[3],
                'categoria_id': result[4],
                'quantidade_em_stock': result[5],
            }
            return produto_data
        else:
            return None
        


def carrinho(request):
    # Get the cart from the session
    cart = request.session.get('cart', {})

    # Calculate the total for the cart and subtotals for each item
    for item in cart.values():
        item['subtotal'] = item['price'] * item['quantity']
    
    # Calculate the overall total
    carrinho_total = sum(item['subtotal'] for item in cart.values())

    # Pass cart and total to the template
    return render(request, 'carrinho.html', {
        'carrinho': cart.items(),
        'carrinho_total': carrinho_total
    })


def add_to_cart(request, produto_id):
    # Get the product data from the stored procedure
    produto_data = get_product_data(produto_id)
    
    if produto_data is None:
        return JsonResponse({'error': 'Produto não encontrado'}, status=404)
    
    # Get quantity from POST (default to 1 if not provided)
    quantity = request.POST.get('quantity', 1)
    
    # Initialize the cart if it's not already in the session
    if 'cart' not in request.session:
        request.session['cart'] = {}
    
    cart = request.session['cart']

    produto_id = str(produto_id)
    
    # If the item is already in the cart, update the quantity
    if produto_id in cart:
        cart[produto_id]['quantity'] += int(quantity)
    else:
        cart[produto_id] = {
            'name': produto_data['nome'],
            'price': float(produto_data['preco']),  # Convert Decimal to float
            'quantity': int(quantity)
        }
    
    # Save the session
    request.session.modified = True
    
    # Return the product details in the response
    return JsonResponse({"success": True})


def finalizar_compra(request):
    if request.method == 'POST':
        try:
            # Recuperar dados do utilizador e do carrinho da sessão
            utilizador_id = request.session.get('utilizador_id')
            cart = request.session.get('cart', {})
            
            if not cart:
                return JsonResponse({'status': 'error', 'message': 'Carrinho vazio.'}, status=400)

            # Dados do formulário (morada)
            morada = request.POST.get('morada')
            estado = 'pendente'  # Estado inicial da encomenda

            # Converter o carrinho no formato esperado
            itens = [
                {'produto_id': int(produto_id), 'quantidade': item['quantity']}
                for produto_id, item in cart.items()
            ]

            # Chamar a função no banco de dados
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT sp_Encomenda_Com_Itens_CREATE(%s, %s, %s, %s);
                """, [utilizador_id, morada, estado, json.dumps(itens)])

            # Limpar o carrinho após finalizar a compra
            request.session['cart'] = {}

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": "Invalid request"}, status=500)
    return JsonResponse({"success": False, "error": "Invalid request"}, status=405)

def dashboard_clientes(request):
    utilizadores = UtilizadorView.objects.all()
    return render(request, 'dashboard_clientes.html', {'utilizadores': utilizadores})

def dashboard_fornecedores(request):
    fornecedores = FornecedorView.objects.all()
    return render(request, 'dashboard_fornecedores.html', {'fornecedores': fornecedores})

def gerar_senha_aleatoria(length=8):
    """ Gera uma senha aleatória segura. """
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(length))

def recuperar_senha(request):
    if request.method == "POST":
        email = request.POST.get("email").strip().lower()

        if not email:
            messages.error(request, "Por favor, insira um email válido.")
            return redirect("recuperar")

        with connection.cursor() as cursor:
            cursor.execute("SELECT utilizador_id FROM utilizador WHERE LOWER(email) = %s", [email])
            user = cursor.fetchone()

        if user:
            nova_senha = gerar_senha_aleatoria()
            senha_hash = hash_password(nova_senha)

            with connection.cursor() as cursor:
                cursor.execute("UPDATE utilizador SET senha = %s WHERE utilizador_id = %s", [senha_hash, user[0]])

            try:
                send_mail(
                    "Recuperação de Senha",
                    f"Sua nova palavra-passe temporária é: {nova_senha}. Por favor, altere-a no próximo login.",
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )

                # Guarda na sessão que este utilizador precisa alterar a senha no próximo login
                request.session["forcar_redefinir_senha"] = True
                request.session["utilizador_id_reset"] = user[0]

                request.session["recuperacao_sucesso"] = "Uma nova palavra-passe foi enviada para o seu email."
                return redirect("recuperar")

            except Exception as e:
                messages.error(request, f"Erro ao enviar email: {str(e)}")

        else:
            messages.error(request, "Email não encontrado na base de dados.")

    sucesso_msg = request.session.pop("recuperacao_sucesso", None)
    return render(request, "recuperar_pass.html", {"sucesso_msg": sucesso_msg})

def alterar_senha(request):
    if "utilizador_id" not in request.session or "forcar_redefinir_senha" not in request.session:
        return redirect("index")  # Se o utilizador não precisar redefinir a senha, redireciona para a home

    if request.method == "POST":
        nova_senha = request.POST.get("nova_senha")
        confirmacao_senha = request.POST.get("confirmacao_senha")

        if nova_senha != confirmacao_senha:
            messages.error(request, "As palavras-passe não coincidem.")
        else:
            senha_hash = hash_password(nova_senha)
            utilizador_id = request.session["utilizador_id"]

            with connection.cursor() as cursor:
                cursor.execute("UPDATE utilizador SET senha = %s WHERE utilizador_id = %s", [senha_hash, utilizador_id])

            # Remover flag da sessão para que não seja pedido novamente
            del request.session["forcar_redefinir_senha"]
            del request.session["utilizador_id_reset"]

            messages.success(request, "Palavra-passe alterada com sucesso!")
            return redirect("index")

    return render(request, "alterar_senha.html")


def produtos_4recentes():
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM ultimos_produtos_adicionados()")
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados

def produtos_nunca_vendidos(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM produtos_nunca_vendidos")
        produtos = cursor.fetchall()

    # Se a view tiver colunas fixas, podes formatar os dados assim:
    produtos_lista = [{'produto_id': row[0], 'nome': row[1]} for row in produtos]

    return JsonResponse(produtos_lista, safe=False)

def top_5_produtos_mais_vendidos(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM top_5_produtos_mais_vendidos")
        produtos = cursor.fetchall()

    produtos_lista = [{'produto_id': row[0], 'nome': row[1], 'quantidade_total_vendida': row[2]} for row in produtos]

    return JsonResponse(produtos_lista, safe=False)

def top_clientes(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM top_clientes")
        clientes = cursor.fetchall()

    clientes_lista = [{'utilizador_id': row[0], 'nome': row[1], 'total_encomendas': row[2]} for row in clientes]

    return JsonResponse(clientes_lista, safe=False)

def categorias_mais_vendidas(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM categorias_mais_vendidas")
        categorias = cursor.fetchall()

    categorias_lista = [{'categoria_id': row[0], 'categoria': row[1], 'total_vendido': row[2]} for row in categorias]

    return JsonResponse(categorias_lista, safe=False)


def produtos_mais_stock():
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM get_top_4_produtos_por_stock()")
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados

def get_all_produtos():
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM get_all_produtos()")
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados

def get_all_categories():
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM get_all_categories()")
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados

def obter_produtos_categoria(id):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM obter_produtos_categoria(%s)", [id])
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados


def procurar_produto_nome(texto):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM procurar_produto_por_nome(%s)", [texto])
        colunas = [col[0] for col in cursor.description]
        resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    return resultados

def get_categorias():
    # Calling the stored procedure using raw SQL
    with connection.cursor() as cursor:
        cursor.callproc('todas_categorias')  # Calling the stored procedure
        rows = cursor.fetchall()  # Fetch all rows returned by the procedure
    
    # Prepare the categories data
    categorias = [{'categoria_id': row[0], 'nome': row[1]} for row in rows]
    
    return categorias

def get_fornecedores():
    # Calling the stored procedure using raw SQL
    with connection.cursor() as cursor:
        cursor.callproc('get_all_fornecedores')  # Calling the stored procedure
        rows = cursor.fetchall()  # Fetch all rows returned by the procedure
    
    # Prepare the categories data
    fornecedores = [{'fornecedor_id': row[0], 'nome': row[1]} for row in rows]
    
    return fornecedores

def obter_encomendas():
    encomendas = EncomendaView.objects.all()
    
    # Dicionário para armazenar as encomendas agrupadas
    grouped_encomendas = defaultdict(lambda: {
        "encomenda_id": None,
        "morada": None,
        "data_encomenda": None,
        "estado": None,
        "produto": [],
        "preco_total": 0,
        "nome_user": None
    })
    
    # Agrupar as encomendas e calcular o total
    for encomenda in encomendas:
        encomenda_id = encomenda.encomenda_id
        encomenda.nome_user = encomenda.nome_user
        item_data = {
            "produto_id": encomenda.produto_id,
            "nome": encomenda.nome_produto,
            "descricao": encomenda.descricao,
            "preco_unitario": encomenda.preco_unitario,
            "quantidade": encomenda.quantidade,
            "preco_total": encomenda.preco_total,
        }
        
        # Adiciona o item à encomenda correspondente
        grouped_encomendas[encomenda_id]["produto"].append(item_data)
        grouped_encomendas[encomenda_id]["preco_total"] += encomenda.preco_total
        
        # A primeira vez que encontramos uma encomenda, salvamos os outros dados
        if grouped_encomendas[encomenda_id]["encomenda_id"] is None:
            grouped_encomendas[encomenda_id]["encomenda_id"] = encomenda.encomenda_id
            grouped_encomendas[encomenda_id]["nome_user"] = encomenda.nome_user
            grouped_encomendas[encomenda_id]["morada"] = encomenda.morada
            grouped_encomendas[encomenda_id]["data_encomenda"] = encomenda.data_encomenda
            grouped_encomendas[encomenda_id]["estado"] = encomenda.estado
    
    # Transformar o dicionário em uma lista para retorno
    data = list(grouped_encomendas.values())
    
    return data

def exportar_relatorio(request):
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if not data_inicio or not data_fim:
        return HttpResponse("Datas inválidas", status=400)

    encomendas = EncomendaView.objects.filter(data_encomenda__range=[data_inicio, data_fim])

    # Separar encomendas por estado
    enviadas = []
    pendentes = []

    for encomenda in encomendas:
        data_formatada = encomenda.data_encomenda.strftime('%Y-%m-%d')
        item = {
            "Encomenda ID": encomenda.encomenda_id,
            "Nome do Cliente": encomenda.nome_user,
            "Morada": encomenda.morada,
            "Data da Encomenda": data_formatada,
            "Estado": encomenda.estado,
            "Produto": encomenda.nome_produto,
            "Preço Total": encomenda.preco_total
        }

        if encomenda.estado.lower() == "enviada":
            enviadas.append(item)
        else:
            pendentes.append(item)

    # Criar ficheiro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Adicionar Título
    titulo = f"Relatório de Vendas - {data_inicio} a {data_fim}"
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Criar função para adicionar tabelas formatadas
    def adicionar_tabela(nome, dados, linha_inicial):
        ws.append([])
        ws.append([nome])
        ws["A{}".format(linha_inicial + 1)].font = Font(bold=True, size=12)

        if dados:
            df = pd.DataFrame(dados)

            # Adicionar cabeçalhos com formatação
            ws.append(df.columns.tolist())
            ultima_linha = ws.max_row
            for col_num, col_nome in enumerate(df.columns, start=1):
                cell = ws.cell(row=ultima_linha, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Adicionar os dados
            for r in df.itertuples(index=False, name=None):
                ws.append(r)

            # Adicionar somatório final
            total = sum(item["Preço Total"] for item in dados)
            ws.append([])
            ws.append(["Total", "", "", "", "", "", total])
            ws["G{}".format(ws.max_row)].font = Font(bold=True)

            # Ajustar a largura das colunas automaticamente
            for col in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                for cell in ws[col_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Adiciona um pouco de espaço extra

        else:
            ws.append(["Sem dados"])

    # Adicionar tabela "Enviadas"
    adicionar_tabela("📦 Encomendas Enviadas", enviadas, ws.max_row + 2)

    # Adicionar tabela "Pendentes"
    adicionar_tabela("⌛ Encomendas Pendentes", pendentes, ws.max_row + 2)

    # Enviar ficheiro como resposta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="relatorio_vendas {data_inicio} a {data_fim}.xlsx"'
    wb.save(response)

    return response


def add_fornecedor(request):

    nome = request.POST['nome']
    contacto = request.POST['contacto']
    endereco = request.POST['endereco']
    with connection.cursor() as cursor:
            cursor.callproc('sp_Fornecedor_CREATE', [nome, contacto, endereco])
            messages.success(request, "Fornecedor registado com sucesso!")
            return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def update_fornecedor(request):
    if request.method == 'POST':
        fornecedor_id = request.POST['fornecedor_id']
        nome = request.POST['nome']
        contacto = request.POST['contacto']
        endereco = request.POST['endereco']

        try:
            with connection.cursor() as cursor:
                cursor.callproc('sp_Fornecedor_UPDATE', [fornecedor_id, nome, contacto, endereco])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def delete_fornecedor(request):
    if request.method == 'POST':
        try:
            print(request.POST) 
            fornecedor_id = request.POST['fornecedor_id']

            with connection.cursor() as cursor:
                cursor.callproc('sp_Fornecedor_DELETE', [fornecedor_id])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def produtos_fornecedor(request, fornecedor_id):
    with connection.cursor() as cursor:
        
        # Chamar a função para obter os produtos
        cursor.execute("SELECT * FROM obter_produtos_por_fornecedor(%s);", [fornecedor_id])
        produtos = cursor.fetchall()

    produtos_data = [
        {"produto_id": p[0], "nome": p[1], "quantidade_em_stock": p[4]}
        for p in produtos
    ]

    return JsonResponse({"produtos": produtos_data})

def delete_cliente(request):
    if request.method == 'POST':
        try:
            print(request.POST) 
            cliente_id = request.POST['cliente_id']

            if not cliente_id:
                return JsonResponse({"success": False, "error": "Cliente ID não fornecido"}, status=400)

            with connection.cursor() as cursor:
                cursor.callproc('sp_Utilizador_DELETE', [cliente_id])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def delete_produto(request):
    if request.method == 'POST':
        try:
            print(request.POST) 
            produto_id = request.POST['produto_id']

            if not produto_id:
                return JsonResponse({"success": False, "error": "Produto ID não fornecido"}, status=400)

            with connection.cursor() as cursor:
                cursor.callproc('sp_Produto_DELETE', [produto_id])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def ver_encomendas_clientes(request, utilizador_id):
    encomendas = obter_encomendas_utilizador(utilizador_id)  # Function to get orders by user
    return render(request, 'encomendas_table.html', {'encomendas': encomendas})

def enviar_encomenda(request, encomenda_id):
    if request.method == 'POST':

        try:
            with connection.cursor() as cursor:
                cursor.callproc('sp_Encomenda_UPDATE', [encomenda_id, 'enviada'])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

def requerir_produto(request):
            data = json.loads(request.body)
            produto_id = data.get('produto_id')
            quantidade = data.get('quantidade')
            preco_str = data.get('preco')
            preco = preco_str.replace(",", ".")
            preco = (float(preco) * int(quantidade)) * 0.75 #Vamos assumir que os produtos são vendidos todos com uma margem de 25% de lucro
            horas_adicionais = random.randint(12, 24)
            data_rececao = datetime.now() + timedelta(hours=horas_adicionais)
            data_rececao_str = data_rececao.strftime('%Y-%m-%d %H:%M:%S')
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM obter_fornecedor_produto(%s)", [produto_id])
                result = cursor.fetchone()  # Retorna apenas um fornecedor
                cursor.callproc('sp_RequisicaoProduto_CREATE', [result[0], produto_id, quantidade, data_rececao_str])
                cursor.callproc('sp_FaturaFornecedor_CREATE', [result[0], preco])

            messages.success(request, "Stock adicionado com sucesso!")

            return JsonResponse({"success": True})

def add_produto(request):
    # Gerar hora random para a receção do produto
    horas_adicionais = random.randint(12, 24)
    data_rececao = datetime.now() + timedelta(hours=horas_adicionais)
    data_rececao_str = data_rececao.strftime('%Y-%m-%d %H:%M:%S')

    nome = request.POST['nome']
    descricao = request.POST.get('descricao', '')
    preco_str = request.POST['preco']
    categoria_id = request.POST['categoria']
    fornecedor_id = request.POST['fornecedor']
    quantidade = request.POST['quantidade']
    preco = preco_str.replace(",", ".")
    preco_fatura = (float(preco) * int(quantidade)) * 0.75 #Vamos assumir que os produtos são vendidos todos com uma margem de 25% de lucro

    # Use a stored procedure to insert the product data (replace with actual procedure if needed)
    with connection.cursor() as cursor:
        cursor.callproc('sp_Produto_CREATE', [nome, descricao, preco, categoria_id, 0])
        produto_id = cursor.fetchone()[0]
        cursor.callproc('sp_RequisicaoProduto_CREATE', [fornecedor_id, produto_id, quantidade, data_rececao_str])
        cursor.callproc('sp_FaturaFornecedor_CREATE', [fornecedor_id, preco_fatura])

    # Send success message
    messages.success(request, "Produto adicionado com sucesso!")

    return JsonResponse({"success": True})


def update_produto(request):
    if request.method == 'POST':
        try:
            # Obtém os dados do formulário
            produto_id = int(request.POST['produto_id'])
            nome = request.POST['nome']
            descricao = request.POST['descricao']
            preco = Decimal(request.POST['preco'])  # Certifica-te de que o preço é convertido para decimal
            categoria_id = int(request.POST['categoria'])
            quantidade_em_stock = int(request.POST['quantidade'])

            # Chama a stored procedure com os dados recebidos
            with connection.cursor() as cursor:
                cursor.callproc('sp_Produto_UPDATE', [
                    produto_id, nome, descricao, preco, categoria_id, quantidade_em_stock
                ])

            return JsonResponse({"success": True})
        
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)



def add_categoria(request):
    # Retrieve data from the POST request
    nome = request.POST['nome']

    # Use a stored procedure to insert the product data (replace with actual procedure if needed)
    with connection.cursor() as cursor:
        cursor.callproc('sp_Categoria_CREATE', [nome])

    # Send success message
    messages.success(request, "Categoria adicionadoa com sucesso!")

    return JsonResponse({"success": True})

    

def update_cliente(request):
    if request.method == 'POST':
        try:
            cliente_id = request.POST['cliente_id']
            nome = request.POST['nome']
            email = request.POST['email']
            isAdmin = request.POST.get('isAdmin', 'false') == 'true'
            senha = request.POST.get('senha', None) 

            
            with connection.cursor() as cursor:
                cursor.callproc('sp_Utilizador_UPDATE', [cliente_id, nome, email, senha, isAdmin])
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)


def obter_encomendas_por_data(data_inicio, data_fim):
    # Convertendo as datas de string para o formato datetime, se necessário
    data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
    data_fim = datetime.strptime(data_fim, "%Y-%m-%d")
    
    # Obter todas as encomendas
    encomendas = EncomendaView.objects.all()
    
    # Dicionário para armazenar as encomendas agrupadas
    grouped_encomendas = defaultdict(lambda: {
        "encomenda_id": None,
        "morada": None,
        "data_encomenda": None,
        "estado": None,
        "produto": [],
        "preco_total": 0,
        "nome_user": None
    })
    
    # Filtrar as encomendas com base na data
    for encomenda in encomendas:
        # Verificar se a encomenda está dentro do intervalo de datas
        if data_inicio <= encomenda.data_encomenda <= data_fim:
            encomenda_id = encomenda.encomenda_id
            item_data = {
                "produto_id": encomenda.produto_id,
                "nome": encomenda.nome_produto,
                "descricao": encomenda.descricao,
                "preco_unitario": encomenda.preco_unitario,
                "quantidade": encomenda.quantidade,
                "preco_total": encomenda.preco_total,
            }
            
            # Adiciona o item à encomenda correspondente
            grouped_encomendas[encomenda_id]["produto"].append(item_data)
            grouped_encomendas[encomenda_id]["preco_total"] += encomenda.preco_total
            
            # A primeira vez que encontramos uma encomenda, salvamos os outros dados
            if grouped_encomendas[encomenda_id]["encomenda_id"] is None:
                grouped_encomendas[encomenda_id]["encomenda_id"] = encomenda.encomenda_id
                grouped_encomendas[encomenda_id]["nome_user"] = encomenda.nome_user
                grouped_encomendas[encomenda_id]["morada"] = encomenda.morada
                grouped_encomendas[encomenda_id]["data_encomenda"] = encomenda.data_encomenda
                grouped_encomendas[encomenda_id]["estado"] = encomenda.estado
    
    # Transformar o dicionário em uma lista para retorno
    data = list(grouped_encomendas.values())
    
    return data


def obter_encomendas_utilizador(utilizador_id):
    # Filtra as encomendas da view com base no utilizador_id
    encomendas = EncomendaView.objects.filter(utilizador_id=utilizador_id)
    
    # Dicionário para armazenar as encomendas agrupadas
    grouped_encomendas = defaultdict(lambda: {
        "encomenda_id": None,
        "morada": None,
        "data_encomenda": None,
        "estado": None,
        "produto": [],
        "preco_total": 0,
    })
    
    # Agrupar as encomendas e calcular o total
    for encomenda in encomendas:
        encomenda_id = encomenda.encomenda_id
        item_data = {
            "produto_id": encomenda.produto_id,
            "nome": encomenda.nome_produto,
            "descricao": encomenda.descricao,
            "preco_unitario": encomenda.preco_unitario,
            "quantidade": encomenda.quantidade,
            "preco_total": encomenda.preco_total,
        }
        
        # Adiciona o item à encomenda correspondente
        grouped_encomendas[encomenda_id]["produto"].append(item_data)
        grouped_encomendas[encomenda_id]["preco_total"] += encomenda.preco_total
        
        # A primeira vez que encontramos uma encomenda, salvamos os outros dados
        if grouped_encomendas[encomenda_id]["encomenda_id"] is None:
            grouped_encomendas[encomenda_id]["encomenda_id"] = encomenda.encomenda_id
            grouped_encomendas[encomenda_id]["morada"] = encomenda.morada
            grouped_encomendas[encomenda_id]["data_encomenda"] = encomenda.data_encomenda
            grouped_encomendas[encomenda_id]["estado"] = encomenda.estado
    
    # Transformar o dicionário em uma lista para retorno
    data = list(grouped_encomendas.values())
    
    return data

def obter_encomenda(encomenda_id):
    encomendas = EncomendaView.objects.filter(encomenda_id=encomenda_id)
    
    # Dicionário para armazenar as encomendas agrupadas
    grouped_encomendas = defaultdict(lambda: {
        "encomenda_id": None,
        "morada": None,
        "data_encomenda": None,
        "estado": None,
        "produto": [],
        "preco_total": 0,
        "nome_user": None,
        "utilizador_id": 0
    })
    
    # Agrupar as encomendas e calcular o total
    for encomenda in encomendas:
        encomenda_id = encomenda.encomenda_id
        item_data = {
            "produto_id": encomenda.produto_id,
            "nome": encomenda.nome_produto,
            "descricao": encomenda.descricao,
            "preco_unitario": encomenda.preco_unitario,
            "quantidade": encomenda.quantidade,
            "preco_total": encomenda.preco_total,
            "utilizador_id": encomenda.utilizador_id
        }
        
        # Adiciona o item à encomenda correspondente
        grouped_encomendas[encomenda_id]["produto"].append(item_data)
        grouped_encomendas[encomenda_id]["preco_total"] += encomenda.preco_total
        
        # A primeira vez que encontramos uma encomenda, salvamos os outros dados
        if grouped_encomendas[encomenda_id]["encomenda_id"] is None:
            grouped_encomendas[encomenda_id]["encomenda_id"] = encomenda.encomenda_id
            grouped_encomendas[encomenda_id]["nome_user"] = encomenda.nome_user
            grouped_encomendas[encomenda_id]["morada"] = encomenda.morada
            grouped_encomendas[encomenda_id]["data_encomenda"] = encomenda.data_encomenda
            grouped_encomendas[encomenda_id]["estado"] = encomenda.estado
            grouped_encomendas[encomenda_id]["utilizador_id"] = encomenda.utilizador_id
    
    # Transformar o dicionário em uma lista para retorno
    data = list(grouped_encomendas.values())
    
    return data

def gerar_pdf_encomenda(request, encomenda_id):
    encomenda_list = obter_encomenda(encomenda_id)  # Obtenção dos dados da encomenda
    encomenda = encomenda_list[0] if encomenda_list else None  # Apenas uma encomenda

    # Caso a encomenda não exista
    if encomenda is None:
        return HttpResponse("Encomenda não encontrada", status=404)

    # Preparando a resposta HTTP com o tipo de conteúdo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fatura_{encomenda["encomenda_id"]}.pdf"'  # Acessando o valor do dicionário

    # Criando o buffer para o PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    # Definindo margens
    margem_esquerda = 50
    margem_topo = 750

    # Adicionando o nome da empresa no topo
    p.setFont("Helvetica-Bold", 18)
    p.drawString(margem_esquerda, margem_topo, "A Minha Loja")
    p.setFont("Helvetica", 10)
    p.drawString(margem_esquerda, margem_topo - 20, "Rua do Fundo, Viseu")
    p.drawString(margem_esquerda, margem_topo - 35, "Telefone: +351 123 456 789 | Email: info@vendas.com")
    
    # Adicionando título "Fatura"
    p.setFont("Helvetica-Bold", 14)
    p.drawString(400, margem_topo, f"Fatura # {encomenda['encomenda_id']}")

    # Adicionando dados da encomenda
    p.setFont("Helvetica", 10)
    p.drawString(50, margem_topo - 70, f"Fatura # {encomenda['encomenda_id']}")
    p.drawString(50, margem_topo - 85, f"Data: {encomenda['data_encomenda'].strftime('%d/%m/%Y')}")
    p.drawString(50, margem_topo - 100, f"Cliente: {encomenda['nome_user']}")
    p.drawString(50, margem_topo - 115, f"Morada: {encomenda['morada']}")

    # Adicionando tabela de produtos
    margem_topo_produtos = margem_topo - 160
    p.setFont("Helvetica-Bold", 10)
    p.drawString(margem_esquerda, margem_topo_produtos, "Produto")
    p.drawString(250, margem_topo_produtos, "Quantidade")
    p.drawString(350, margem_topo_produtos, "Preço Unitário")
    p.drawString(450, margem_topo_produtos, "Preço Total")

    # Linha horizontal
    p.setLineWidth(0.5)
    p.line(margem_esquerda, margem_topo_produtos - 5, 550, margem_topo_produtos - 5)

    # Adicionando os itens da encomenda
    y_position = margem_topo_produtos - 20
    p.setFont("Helvetica", 10)
    for item in encomenda['produto']:
        p.drawString(margem_esquerda, y_position, item['nome'])
        p.drawString(250, y_position, str(item['quantidade']))
        p.drawString(350, y_position, f"€{item['preco_unitario']:,.2f}")
        p.drawString(450, y_position, f"€{item['preco_total']:,.2f}")
        y_position -= 20

    # Linha horizontal após os produtos
    p.line(margem_esquerda, y_position, 550, y_position)

    # Total da encomenda
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, y_position - 20, "Total:")
    p.setFont("Helvetica", 12)
    preco_total = f"€{encomenda['preco_total']:,.2f}"
    p.drawString(450, y_position - 20, preco_total)

    # Finalizando o PDF
    p.showPage()
    p.save()

    # Retornando o PDF gerado
    buffer.seek(0)
    response.write(buffer.read())
    return response